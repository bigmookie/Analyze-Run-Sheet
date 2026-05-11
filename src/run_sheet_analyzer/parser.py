"""Parse a run sheet Excel file into normalized RunSheetRow / Tract objects.

Strict header validation: any missing required column raises MissingColumnsError
with the exact list of missing columns. Header matching is case-insensitive,
whitespace-tolerant, and accepts a small set of common synonyms.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


CANONICAL_HEADERS: dict[str, list[str]] = {
    "book":                ["book"],
    "page":                ["page"],
    "date_recorded":       ["date recorded", "recorded"],
    "doc_title":           ["document title", "doc title"],
    "grantor":             ["grantor", "grantor(s)", "grantors"],
    "grantee":             ["grantee", "grantee(s)", "grantees"],
    "brief_description":   ["brief description", "brief legal description", "legal"],
    "tract":               ["tract"],
    "mineral_reservations":["mineral reservations", "reservations"],
    "notes":               ["notes"],
    "exception":           ["exception", "exceptions"],
    "mineral_transfer":    ["mineral transfer", "min transfer"],
}

REQUIRED_FIELDS: set[str] = set(CANONICAL_HEADERS.keys())

OPTIONAL_HEADERS: dict[str, list[str]] = {
    "instrument_no":       ["instrument no", "inst no", "instrument number"],
}


class MissingColumnsError(Exception):
    def __init__(self, missing: list[str]):
        self.missing = missing
        labels = [CANONICAL_HEADERS[k][0].title() for k in missing]
        super().__init__("Missing required column(s): " + ", ".join(labels))


@dataclass
class RunSheetRow:
    row_index: int
    book: str
    page: str
    instrument_no: str | None
    date_recorded: date | None
    doc_title: str
    grantors: list[str]
    grantees: list[str]
    tracts: list[str]        # numeric base tracts the row belongs to; never includes "LE"/"NS"
    is_less_except: bool     # row is an L&E carve-out from each of `tracts`
    is_not_subject: bool     # row is NS
    brief_description: str
    mineral_reservations: str
    notes: str
    exception_flag: bool
    mineral_transfer_flag: bool

    @property
    def cite(self) -> str:
        bits = [f"Book {self.book}, Page {self.page}"]
        if self.instrument_no:
            bits.append(f"Inst. No. {self.instrument_no}")
        return ", ".join(bits)


@dataclass
class Tract:
    id: str
    rows: list[RunSheetRow] = field(default_factory=list)
    le_rows: list[RunSheetRow] = field(default_factory=list)


@dataclass
class ParsedRunSheet:
    path: Path
    rows: list[RunSheetRow]
    tracts: dict[str, Tract]      # by tract id
    not_subject_rows: list[RunSheetRow]
    unparseable_le_rows: list[RunSheetRow]   # bare "LE" with no base tract

    def tract_ids(self) -> list[str]:
        return sorted(self.tracts.keys())


def _norm_header(h: object) -> str:
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).strip().lower())


def _build_header_map(header_row: list[object]) -> tuple[dict[str, int], dict[str, int]]:
    """Return (required_map, optional_map): field-name → 0-based column index."""
    normalized = [_norm_header(c) for c in header_row]
    required: dict[str, int] = {}
    optional: dict[str, int] = {}
    for field_name, synonyms in CANONICAL_HEADERS.items():
        for idx, h in enumerate(normalized):
            if h in synonyms:
                required[field_name] = idx
                break
    for field_name, synonyms in OPTIONAL_HEADERS.items():
        for idx, h in enumerate(normalized):
            if h in synonyms:
                optional[field_name] = idx
                break
    return required, optional


def _cell(row: tuple, idx: int | None) -> object:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _split_pipe(v: object) -> list[str]:
    s = _str(v)
    if not s:
        return []
    return [p.strip() for p in s.split("|") if p.strip()]


def _bool_flag(v: object) -> bool:
    s = _str(v).lower()
    return s not in ("", "0", "false", "no")


_DATE_FORMATS = (
    "%m/%d/%Y",
    "%Y-%m-%d",
    "%m-%d-%Y",
    "%Y/%m/%d",
    "%m/%d/%y",
)


def _parse_date(v: object) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _str(v)
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_tract_tokens(v: object) -> tuple[list[str], bool, bool]:
    """Return (base_tracts, is_less_except, is_not_subject).

    Rules:
      - "NS" anywhere → not subject.
      - First token "LE" → L&E carve-out from the remaining tokens (must have ≥1).
      - "LE" appearing elsewhere is treated the same as first-position LE.
      - "LE" with no base tract → unparseable_le (caller decides what to do).
      - Otherwise the tokens are base tracts the row belongs to.
    """
    s = _str(v)
    if not s:
        return [], False, False
    tokens = [t.strip() for t in s.split(",") if t.strip()]
    if not tokens:
        return [], False, False
    upper = [t.upper() for t in tokens]
    if "NS" in upper:
        return [], False, True
    if "LE" in upper:
        base = [t for t in tokens if t.upper() != "LE"]
        return base, True, False
    return tokens, False, False


def parse(path: str | Path) -> ParsedRunSheet:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row: first row with any non-empty cell.
    header_row_idx = 0
    header_row: list[object] = []
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(c is not None and str(c).strip() for c in row):
            header_row_idx = r
            header_row = row
            break
    if not header_row:
        raise ValueError("Run sheet appears to be empty.")

    req, opt = _build_header_map(header_row)
    missing = sorted(REQUIRED_FIELDS - req.keys())
    if missing:
        raise MissingColumnsError(missing)

    rows: list[RunSheetRow] = []
    tracts: dict[str, Tract] = {}
    not_subject: list[RunSheetRow] = []
    unparseable_le: list[RunSheetRow] = []

    for r in range(header_row_idx + 1, ws.max_row + 1):
        raw = tuple(ws.cell(r, c).value for c in range(1, ws.max_column + 1))
        if not any(_str(c) for c in raw):
            continue

        tract_tokens, is_le, is_ns = _parse_tract_tokens(_cell(raw, req["tract"]))

        row_obj = RunSheetRow(
            row_index=r,
            book=_str(_cell(raw, req["book"])),
            page=_str(_cell(raw, req["page"])),
            instrument_no=_str(_cell(raw, opt["instrument_no"])) or None if "instrument_no" in opt else None,
            date_recorded=_parse_date(_cell(raw, req["date_recorded"])),
            doc_title=_str(_cell(raw, req["doc_title"])),
            grantors=_split_pipe(_cell(raw, req["grantor"])),
            grantees=_split_pipe(_cell(raw, req["grantee"])),
            tracts=tract_tokens,
            is_less_except=is_le,
            is_not_subject=is_ns,
            brief_description=_str(_cell(raw, req["brief_description"])),
            mineral_reservations=_str(_cell(raw, req["mineral_reservations"])),
            notes=_str(_cell(raw, req["notes"])),
            exception_flag=_bool_flag(_cell(raw, req["exception"])),
            mineral_transfer_flag=_bool_flag(_cell(raw, req["mineral_transfer"])),
        )
        rows.append(row_obj)

        if row_obj.is_not_subject:
            not_subject.append(row_obj)
            continue

        if row_obj.is_less_except:
            if not tract_tokens:
                unparseable_le.append(row_obj)
                continue
            for t in tract_tokens:
                tracts.setdefault(t, Tract(id=t)).le_rows.append(row_obj)
            continue

        for t in tract_tokens:
            tracts.setdefault(t, Tract(id=t)).rows.append(row_obj)

    return ParsedRunSheet(
        path=path,
        rows=rows,
        tracts=tracts,
        not_subject_rows=not_subject,
        unparseable_le_rows=unparseable_le,
    )
