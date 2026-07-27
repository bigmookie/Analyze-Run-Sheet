"""Assign run-sheet rows to tract units when the Tract column is blank.

The examiner defines the tracts — either by typing units (section /
quarter / quarter-quarter / full legal description) or by providing a legal
description document (.txt / .docx) that names tracts and gives each a full legal
description. Claude then reads each run-sheet row's brief description and decides
which tract(s) the instrument affects. We re-bucket the rows into
``parsed.tracts`` exactly as the parser would, so the rest of the pipeline is
untouched.

One Claude call to extract tracts from a document, one to match rows. No PLSS
parsing in Python — the model does the aliquot reasoning.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass
from importlib import resources
from pathlib import Path
from typing import Callable

import anthropic
import openpyxl

from run_sheet_analyzer import analyzer
from run_sheet_analyzer.parser import (
    ParsedRunSheet,
    RunSheetRow,
    Tract,
    _build_header_map,
)


@dataclass
class TractUnit:
    """A tract the rows get matched against.

    ``id`` is the short label used as a dict key / cache filename / report
    heading (e.g. "Tract 1"). ``description`` is the full legal description used
    as matching context (for a typed unit, it's just the unit text itself).
    """
    id: str
    description: str


def _safe_id(unit: str) -> str:
    """Make a label usable as a dict key and cache filename.

    The cache writes ``out/<tract_id>.txt`` (see cache.py), so the id must not
    contain path separators. We keep it human-readable otherwise.
    """
    s = re.sub(r"\s+", " ", unit.strip())
    return s.replace("/", "-").replace("\\", "-")


def _load_prompt(name: str) -> str:
    return resources.files("run_sheet_analyzer.prompts").joinpath(name).read_text(encoding="utf-8")


def _format_row(row: RunSheetRow) -> str:
    desc = (row.brief_description or "—").replace("\n", "; ")
    return f"- row {row.row_index}: [{row.doc_title or '—'}] {desc}"


def _extract_json(text: str) -> dict:
    """Tolerantly pull the JSON object out of the model's reply."""
    s = text.strip()
    # Strip ``` / ```json fences if present.
    if s.startswith("```"):
        s = re.sub(r"^```[a-zA-Z]*\n?", "", s)
        s = re.sub(r"\n?```$", "", s).strip()
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        # Fall back to the first {...} span.
        start = s.find("{")
        end = s.rfind("}")
        if start != -1 and end > start:
            return json.loads(s[start : end + 1])
        raise


# Rows are matched in batches so the JSON response never approaches the model's
# output cap (a single call over a 193-row sheet truncated mid-JSON).
ASSIGN_BATCH_SIZE = 40


def _message(client: anthropic.Anthropic, prompt: str, log: Callable[[str], None]):
    return analyzer._create_with_retry(
        client,
        kwargs=dict(
            model=analyzer.active_model(client),
            max_tokens=analyzer.MAX_TOKENS,
            messages=[{"role": "user", "content": prompt}],
        ),
        on_progress=log,
    )


def _text(resp) -> str:
    return "".join(b.text for b in resp.content if getattr(b, "type", None) == "text")


def _message_to_completion(
    client: anthropic.Anthropic, prompt: str, log: Callable[[str], None]
) -> tuple[str, bool]:
    """Run one message to completion, auto-continuing if it hits the output cap.

    Mirrors ``analyzer._generate``'s continuation loop for the plain (no system,
    no thinking) message shape used here. Concatenating the raw text chunks is
    seamless — each continuation resumes exactly where the previous one stopped,
    so a JSON reply split across the cap reassembles into valid JSON. Returns
    ``(accumulated_text, truncated)``; ``truncated`` is True only if the model
    was still cut off after ``MAX_CONTINUATIONS`` attempts.
    """
    messages: list[dict] = [{"role": "user", "content": prompt}]
    accumulated = ""
    for attempt in range(1 + analyzer.MAX_CONTINUATIONS):
        if attempt > 0:
            log(f"extraction truncated; continuation #{attempt} …")
            messages = messages + [
                {"role": "assistant", "content": accumulated},
                {"role": "user", "content": analyzer._CONTINUE_PROMPT},
            ]
        resp = analyzer._create_with_retry(
            client,
            kwargs=dict(
                model=analyzer.active_model(client),
                max_tokens=analyzer.MAX_TOKENS,
                messages=messages,
            ),
            on_progress=log,
        )
        accumulated += _text(resp)
        if resp.stop_reason != "max_tokens":
            return accumulated, False
    return accumulated, True


# ──────────────────────────────────────────────────────────────────────────
# Reading tracts from a legal-description document
# ──────────────────────────────────────────────────────────────────────────


def read_document(path: str | Path) -> str:
    """Return the plain text of a .txt or .docx legal-description file."""
    p = Path(path)
    if p.suffix.lower() == ".docx":
        from docx import Document  # python-docx; already a project dependency
        return "\n".join(par.text for par in Document(str(p)).paragraphs)
    return p.read_text(encoding="utf-8", errors="replace")


def extract_units(
    *,
    client: anthropic.Anthropic,
    document_text: str,
    on_progress: Callable[[str], None] | None = None,
) -> list[TractUnit]:
    """Use Claude to pull named tracts + their legal descriptions from a document."""
    log = on_progress or (lambda s: None)
    log(f"reading tracts from description document via {analyzer.active_model(client)} …")
    prompt = _load_prompt("extract.md").replace("{document}", document_text)
    text, truncated = _message_to_completion(client, prompt, log)
    if truncated:
        raise ValueError(
            "tract-extraction response was still truncated at the model output cap "
            f"after {analyzer.MAX_CONTINUATIONS} continuations — the legal-description "
            "document may be too large; split it into smaller files and retry."
        )
    try:
        data = _extract_json(text)
    except json.JSONDecodeError as e:
        raise ValueError(f"Could not parse the tract-extraction response as JSON: {e}") from e

    units: list[TractUnit] = []
    seen: set[str] = set()
    for t in data.get("tracts", []):
        label = str(t.get("id", "")).strip()
        desc = str(t.get("description", "")).strip()
        if not label:
            continue
        tid = _safe_id(label)
        if tid in seen:
            continue
        seen.add(tid)
        units.append(TractUnit(id=tid, description=desc or label))
    log(f"extracted {len(units)} tract(s) from the document.")
    return units


# ──────────────────────────────────────────────────────────────────────────
# Matching rows to tracts
# ──────────────────────────────────────────────────────────────────────────


def propose_assignment(
    *,
    client: anthropic.Anthropic,
    rows: list[RunSheetRow],
    units: list[TractUnit],
    granularity: str,
    on_progress: Callable[[str], None] | None = None,
) -> dict[int, list[str]]:
    """Ask Claude which tract(s) each row affects.

    Returns a mapping of row_index → list of matched tract ids. Rows that matched
    no unit are absent / map to an empty list; the caller gets the definitive
    unassigned list back from ``apply_assignment``.
    """
    log = on_progress or (lambda s: None)
    template = _load_prompt("assign.md")
    units_text = "\n\n".join(f"### {u.id}\n{u.description}" for u in units)
    valid_units = {u.id for u in units}

    batches = [rows[i : i + ASSIGN_BATCH_SIZE] for i in range(0, len(rows), ASSIGN_BATCH_SIZE)]
    log(f"matching {len(rows)} rows to {len(units)} tract(s) in {len(batches)} batch(es) "
        f"via {analyzer.active_model(client)} …")

    mapping: dict[int, list[str]] = {}
    for bi, batch in enumerate(batches, 1):
        if len(batches) > 1:
            log(f"batch {bi}/{len(batches)} ({len(batch)} rows) …")
        rows_text = "\n".join(_format_row(r) for r in batch)
        prompt = template.format(granularity=granularity, units=units_text, rows=rows_text)

        resp = _message(client, prompt, log)
        if resp.stop_reason == "max_tokens":
            raise ValueError(
                "assignment response hit the output cap before completing — "
                "reduce ASSIGN_BATCH_SIZE and retry."
            )
        try:
            data = _extract_json(_text(resp))
        except json.JSONDecodeError as e:
            raise ValueError(f"Could not parse the assignment response as JSON: {e}") from e

        batch_rows = {r.row_index for r in batch}
        for entry in data.get("assignments", []):
            try:
                row_idx = int(entry["row"])
            except (KeyError, TypeError, ValueError):
                continue
            if row_idx not in batch_rows:
                continue
            mapping[row_idx] = [u for u in entry.get("units", []) if u in valid_units]

    n_unassigned = sum(1 for r in rows if not mapping.get(r.row_index))
    log(f"assigned {len(rows) - n_unassigned}/{len(rows)} rows; {n_unassigned} need review.")
    return mapping


def apply_assignment(parsed: ParsedRunSheet, mapping: dict[int, list[str]]) -> list[RunSheetRow]:
    """Populate ``parsed.tracts`` from the row→tract mapping.

    Mirrors the parser's bucketing (parser.py): a row matching multiple tracts is
    replicated into each. Returns the rows that matched no tract.
    """
    parsed.tracts.clear()
    unassigned: list[RunSheetRow] = []

    for row in parsed.rows:
        ids = mapping.get(row.row_index) or []
        row.tracts = ids
        if not ids:
            unassigned.append(row)
            continue
        for tid in ids:
            parsed.tracts.setdefault(tid, Tract(id=tid)).rows.append(row)
    return unassigned


def write_assigned_run_sheet(
    src_path: str | Path, parsed: ParsedRunSheet, dest_path: str | Path
) -> Path:
    """Write a copy of the run sheet with the Tract column filled from the
    assignment. The original file is not modified.

    Each row's assigned tract ids are joined with ", " (the run sheet's
    multi-tract convention) and written into the existing Tract column at the
    row's original worksheet position. Unassigned rows are left blank.
    """
    src_path = Path(src_path)
    dest_path = Path(dest_path)

    wb = openpyxl.load_workbook(src_path)  # keep formatting/formulas
    ws = wb[wb.sheetnames[0]]

    # Locate the header row (first non-blank row) and the Tract column.
    header_row = None
    for r in range(1, ws.max_row + 1):
        rowvals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(c is not None and str(c).strip() for c in rowvals):
            header_row = rowvals
            break
    if header_row is None:
        raise ValueError("Run sheet has no header row to locate the Tract column.")

    req, _opt = _build_header_map(header_row)
    if "tract" not in req:
        raise ValueError("Could not find the Tract column to fill.")
    tract_col = req["tract"] + 1  # openpyxl is 1-based

    for row in parsed.rows:
        joined = ", ".join(row.tracts)
        ws.cell(row=row.row_index, column=tract_col).value = joined or None

    dest_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest_path)
    return dest_path
